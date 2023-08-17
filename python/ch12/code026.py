# Copyright © https://steam.oxxostudio.tw

import os
os.chdir('/content/drive/MyDrive/Colab Notebooks')  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl
wb = openpyxl.load_workbook('test.xlsx', data_only=True)  # 設定 data_only=True 只讀取計算後的數值

s1 = wb['工作表1']
s2 = wb['工作表2']

print(s1['A1'].value)        # 取出 A1 的內容
print(s1.cell(1, 1).value)   # 等同取出 A1 的內容
print(s2['B2'].value)        # 取出 B2 的內容
print(s2.cell(2, 2).value)   # 等同取出 B2 的內容

