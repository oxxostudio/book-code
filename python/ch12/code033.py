# Copyright © https://steam.oxxostudio.tw

import os
os.chdir('/content/drive/MyDrive/Colab Notebooks')  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl
wb = openpyxl.load_workbook('oxxo.xlsx', data_only=True)

s1 = wb['工作表1']                        # 開啟工作表 1
s2 = wb['工作表2']                        # 開啟工作表 2
s1.sheet_properties.tabColor = 'ff0000'  # 修改工作表 1 頁籤顏色為紅色
s2.sheet_properties.tabColor = 'ffff00'  # 修改工作表 2 頁籤顏色為黃色

wb.create_sheet("工作表3")      # 插入工作表 3 在最後方
wb.create_sheet("工作表1.5",1)  # 插入工作表 1.5 在第二個位置 ( 工作表 1 和 2 的中間 )
wb.create_sheet("工作表0", 0)   # 插入工作表 0 在第一個位置

wb.copy_worksheet(s2)          # 複製工作表 2 放到最後方

s1.title='oxxo'                # 修改工作表 1 的名稱為 oxxo
s2.title='studio'              # 修改工作表 2 的名稱為 studio

wb.save('test2.xlsx')

