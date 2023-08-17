# Copyright © https://steam.oxxostudio.tw

import os
os.chdir('/content/drive/MyDrive/Colab Notebooks')  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl
from openpyxl.styles import Font, PatternFill       # 載入 Font 和 PatternFill 模組
wb = openpyxl.load_workbook('oxxo.xlsx', data_only=True)

s1 = wb['工作表1']
s1['e1'].font = Font(name='Arial', color='ff0000', size=30, bold=True) # 設定 g1 儲存格的文字樣式
s1['f1'].fill = PatternFill(fill_type="solid", fgColor="DDDDDD")       # 設定 f1 儲存格的背景樣式

wb.save('test2.xlsx')

