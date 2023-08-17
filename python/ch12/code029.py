# Copyright © https://steam.oxxostudio.tw

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

print(column_index_from_string('A'))    # 1
print(column_index_from_string('AA'))   # 27

print(get_column_letter(5))             # E
print(get_column_letter(100))           # CV

