# Copyright © https://steam.oxxostudio.tw

import os
os.chdir('/content/drive/MyDrive/Colab Notebooks')  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl
wb = openpyxl.load_workbook('oxxostudio.xlsx')     # 開啟 Excel 檔案

names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
s1 = wb['工作表1']        # 取得工作表名稱為「工作表1」的內容
s2 = wb.active           # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )

print(names)
# 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數
print(s1.title, s1.max_row, s1.max_column)
print(s2.title, s2.max_row, s2.max_column)

