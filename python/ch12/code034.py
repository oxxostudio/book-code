# Copyright © https://steam.oxxostudio.tw

import os
os.chdir('/content/drive/MyDrive/Colab Notebooks')  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl
wb = openpyxl.load_workbook('oxxo.xlsx', data_only=True)

s1 = wb['工作表1']            # 開啟工作表 1
s1['A1'].value = 'apple'     # 儲存格 A1 內容為 apple
s1['A2'].value = 'orange'    # 儲存格 A2 內容為 orange
s1['A3'].value = 'banana'    # 儲存格 A3 內容為 banana
s1.cell(1,2).value = 100     # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
s1.cell(2,2).value = 200     # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
s1.cell(3,2).value = 300     # 儲存格 B3 內容 ( row=3, column=2 ) 為 300

wb.save('test2.xlsx')

